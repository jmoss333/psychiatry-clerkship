from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook(); ws = wb.active; ws.title = "Master Index"
cols = ["Title","Category","Subcategory","Format","Topic","Educational Level","Difficulty","Est. Time","Clinical Relevance","Current Status","Priority","Location","Recommended Action"]
R = [
["Psychiatry Clerkship Library — Master Plan (this audit)","Planning","Audit/Architecture","MD","Library design","Faculty","—","45m","High","Exists","P0","~/ (home)","Use as front door / README"],
["Education Hub Blueprint (3-path architecture)","Planning","Architecture","MD","Hub design","Faculty","—","20m","Med","Revise","P1","~/Education_Hub_Blueprint.md","Fold into master plan"],
["Library Plan & Audit Roadmap (Book/Podcast)","Planning","Architecture","MD","Media library","Faculty","—","25m","Med","Exists","P3","~/Library_Plan_and_Audit_Roadmap.md","Build Library v1"],
["A Day on the Unit — Walkthrough","Orientation","Logistics","DOCX","Unit orientation","MS3","Intro","20m","High","Revise","P1","iCloud Work&Career / repo","Wrap into 00_START_HERE syllabus"],
["MH Advanced Clerkship — Elective Application AY26-27","Faculty","Accreditation","PDF","Elective design","Faculty","—","30m","High","Exists","P0","~/Downloads","Use to set learning objectives"],
["Landmark Psychiatry Teaching Guide (100 papers)","Evidence & Reading","Landmark Library","MD","Evidence base","MS3→Attending","Mixed","Reference","High","Revise","P0","teaching/landmark-psychiatry-teaching-guide.md","Canonical; reconcile 15-vs-16 count"],
["15/16-paper Reading Pathway (4-wk)","Evidence & Reading","Reading Pathway","MD","Curated readings","MS3","Intro→Adv","~16 papers","High","Revise","P0","(within Landmark guide)","Re-sequence to 6 weeks"],
["Journal Club in a Box — 6 packets","Evidence & Reading","Journal Club","MD","EBM/critical appraisal","MS3→Resident","Inter","45m each","High","Exists","P1","(within Landmark guide)","Schedule across 6 weeks"],
["Teaching Script Template (10-min whiteboard)","Faculty","Teaching Method","MD","Teaching skill","Resident→Faculty","Inter","10m","High","Exists","P1","(within Landmark guide)","Resident overlay"],
["Landmark library — Google Drive export","Evidence & Reading","Landmark Library","Folder","Evidence base","All","Mixed","Reference","High","Merge","P0","Google Drive","Keep as distribution export"],
["Resident Handout (2-page)","Evidence & Reading","Landmark Library","PDF/DOCX","Quick reference","Resident","Inter","10m","High","Exists","P1","Google Drive","Add to Resident track"],
["Facilitator Teaching Script + Answer Keys","Faculty","Journal Club","PDF/DOCX","Facilitation","Faculty","Inter","30m","High","Exists","P1","Google Drive","Pair with JC packets"],
["Landmark Article Library (live)","Evidence & Reading","Landmark Library","Notion DB","Evidence base","All","Mixed","Reference","High","Merge","P0","Notion","Keep as editable live view"],
["ASCP Psychopharmacology Curriculum (11th ed)","Psychopharmacology","Reference Curriculum","Folder","Psychopharm","Resident→Faculty","Adv","Reference","High","Exists","P3","iCloud Gen Psych Resources","Reference for psychopharm primer"],
["Inpatient Guideline Surveillance 2023-2026","Evidence & Reading","Guidelines","DOCX","Guideline tracking","Resident→Faculty","Adv","Reference","High","Exists","P1","~/Inpatient_Psychiatry_Guideline_Surveillance_2023-2026.docx","Guidelines hub"],
["APA Guideline Highlights Plan (+ Build Plan v2)","Evidence & Reading","Guidelines","MD","Guideline highlights","Resident","Inter","20m","Med","Revise","P1","~/APA_Guideline_Highlights_Plan.md","Execute highlights build"],
["Relational Psychiatry Teaching Manual v2","Teaching Spine","RSS Frame","DOCX/MD","Relational psychiatry","MS3→Resident","Inter","Reference","High","Exists","P0","teaching/Relational_Psychiatry_Teaching_Manual_v2.docx","Canonical; archive v1"],
["Composite Teaching Cases v1","Cases & Simulation","Composite Cases","MD","Case-based teaching","MS3→Resident","Inter","30m","High","Revise","P1","teaching/archive/composite_cases_v1.md","Promote from archive; map to weeks"],
["RSS Video Scripts (10)","Media","Video Scripts","MD","RSS layers / family","All","Intro","8m each","Med","Merge","P0","teaching/video-scripts/ (+ dup dir)","Merge two dirs to one"],
["13 NotebookLM Projects + Audio Overviews","Media","Audio/AI Study","MD/TXT/Audio","Mixed topics","All","Mixed","Varies","Med","Exists","P3","teaching/notebooklm-projects/","Index into 12_Media"],
["ED→Inpatient Psych Capstone package","Patient/Family Ed","ED Boarding","Mixed","Crisis/boarding","Patient/Staff","Intro","Reference","Med","Exists","P3","teaching/ed-psych-capstone/","Keep; link staff materials"],
["Gen Psych Didactic Decks (Bootcamp, Barkley, CPT, Exposure)","Psychotherapy","Didactics","PPTX/PDF","Therapy modalities","MS3→Resident","Inter","45m each","High","Exists","P1","iCloud Gen Psych Resources","Harvest into Week 3"],
["Biopsychosocial Formulation Samples","Clinical Skills","Case Formulation","DOCX","Formulation","MS3","Intro","20m","High","Expand","P1","iCloud Gen Psych Resources","Build formulation worksheet"],
["Notion Teaching Curriculum DB + Teaching-Prep Agent","AI & Prompts","Automation","Notion","Teaching ops","Faculty","—","—","Med","Exists","P3","Notion","Wire to curriculum"],
["Mental Status Exam module (interactive + pocket guide)","Clinical Skills","Mental Status Exam","HTML+MD","MSE","MS3","Intro","30m","High","Exists","P1","02_Clinical_Skills/Mental_Status_Exam/","BUILT: interactive tool + Codex pocket guide"],
["Decisional Capacity module (interactive)","Acute & Safety","Decisional Capacity","HTML","Capacity","MS3","Intro","20m","High","Exists","P1","04_Acute_and_Safety/Decisional_Capacity/","BUILT: 4-abilities tool + note generator"],
["Oral Presentation / Rounds module (interactive)","Clinical Skills","Oral Presentations","HTML+MD","Presentations","MS3","Intro","15m","High","Exists","P1","02_Clinical_Skills/Oral_Presentations/","BUILT: timer tool + Codex doc/present guide"],
["Psychiatric Interviewing module","Clinical Skills","Interviewing","—","Interview","MS3","Intro","30m","High","Expand","P1","RSS scripts + decks","Author interviewing module + clips"],
["Documentation exemplars + checklist","Clinical Skills","Documentation","MD","Notes","MS3","Intro","20m","High","Expand","P1","iCloud Clinical Tools (template)","Author note exemplars"],
["Differential Diagnosis / Formulation pocket guide","Clinical Skills","Differential Dx","MD","DDx","MS3","Inter","30m","High","Exists","P1","14_Tracks/MS3/Student_Ready_Pack/02_pocket_guides/","Covered by Codex formulation/DDx guide"],
["Psychiatric documentation template","Clinical Skills","Documentation","MD","Notes","MS3→Resident","Intro","10m","High","Exists","P1","iCloud Clinical Tools","Use as base"],
["BHU2 Top-20 Pocket Card 2026","Clinical Skills","Pocket Reference","HTML","Inpatient quick-ref","MS3→Resident","Intro","Reference","High","Exists","P0","~/BHU2_Top20_Pocket_Card_2026.html","Flagship pocket card"],
["Suicide Safer-Discharge Checklist","Acute & Safety","Suicide Risk","HTML","Risk/disposition","MS3→Resident","Inter","15m","High","Exists","P1","~/BHU2_Suicide_Safer_Discharge_Checklist.html","Link Week 1/6"],
["C-SSRS / Safety-Plan EHR Field Spec","Acute & Safety","Suicide Risk","DOCX","Risk tools","Resident/Faculty","Inter","Reference","High","Exists","P1","~/CSSRS_SafetyPlan_EHR_Field_Spec.docx","Link to safety-plan skill"],
["Delirium Prevention Checklist + Order Set","Acute & Safety","Delirium","HTML/DOCX","Delirium","MS3→Resident","Inter","15m","High","Exists","P1","~/BHU2_Delirium_*","Week 5"],
["Catatonia discharge instructions","Acute & Safety","Catatonia","PDF","Catatonia","MS3→Resident","Inter","10m","High","Exists","P1","~/Documents/.../catatonia dc instruct.pdf","Pair with BFCRS card"],
["Restraint / Physical-Holding Checklist + Order Set","Acute & Safety","Agitation/Restraint","HTML/DOCX","Agitation","Resident/Faculty","Inter","15m","High","Exists","P2","~/BHU2_Restraint_*","Within agitation ladder"],
["Agitation management ladder module","Acute & Safety","Agitation/Restraint","—","Agitation","MS3","Intro","20m","High","Expand","P2","(restraint assets)","Author verbal→PRN→restraint"],
["Withdrawal scales card — CIWA-Ar & COWS (interactive)","Core Topics","SUD/Withdrawal","HTML","Withdrawal","MS3","Intro","10m","High","Exists","P2","03_Core_Topics/SUD_Withdrawal/","BUILT: interactive tally + severity bands"],
["Violence-risk one-pager + Brøset (interactive)","Acute & Safety","Violence Risk","HTML","Violence risk","MS3","Inter","10m","High","Exists","P2","04_Acute_and_Safety/Violence_Risk/","BUILT: factors + BVC tally + management"],
["Benzodiazepine Taper Checklist + Order-Set Spec","Psychopharmacology","Protocol Library","HTML/DOCX","Taper","Resident/Faculty","Inter","Reference","High","Exists","P1","~/BHU2_Benzodiazepine_*","Protocol library"],
["Clozapine Post-REMS Workflow + Order-Set Spec","Psychopharmacology","Protocol Library","HTML/DOCX","Clozapine","Resident/Faculty","Adv","Reference","High","Exists","P1","~/BHU2_Clozapine_*","Protocol library"],
["Student Psychopharm Primer (Top-10 inpatient)","Psychopharmacology","Student Primer","—","Psychopharm","MS3","Intro","30m","High","Expand","P1","(protocols + ASCP)","Author student tier"],
["BHU2 Implementation Packet + Committee Deck","Faculty","Systems/QI","DOCX/PPTX","Implementation","Faculty","Adv","Reference","Med","Exists","P3","~/BHU2_Implementation_*","Faculty/systems track"],
["Family Meeting Science suite (12-cases, sys review, taxonomy, LEFI)","Family & Relational","Research+Teaching","MD/XLSX","Family therapy","Resident→Faculty","Adv","Reference","High","Exists","P0","family-meeting-science/","Core; link Week 4"],
["Family Meeting Playbook (90-min)","Family & Relational","Family Meeting","MD","Family meeting","MS3→Resident","Inter","30m","High","Exists","P1","clinical-materials/Clinical_Implementation_Spine/","Week 4 skill"],
["Family-Therapy Inpatient Didactic deck (CANONICAL)","Family & Relational","Didactic Deck","PPTX","Family therapy","MS3→Resident","Inter","45m","High","Merge","P0","Downloads/iCloud/GDrive (~12 versions)","Pick 1 canonical; archive rest"],
["'The Family is the Milieu' FINAL","Family & Relational","Narrative Deck","DOCX/PPTX","Family therapy","All","Intro","30m","High","Exists","P1","Downloads/iCloud","Canonical narrative"],
["FTM papers/briefs + FTM audio library","Family & Relational","Reference/Audio","PDF/Audio","Family therapy","Resident","Inter","Varies","Med","Exists","P3","_assets/ftm-papers, _assets/ftm-audio","Index into 12_Media"],
["Clinical Implementation Spine (First 3 Sessions, MVP Milieu, Adherence, Discharge)","Family & Relational","Implementation","MD","Milieu/family ops","Resident→Faculty","Adv","Reference","High","Exists","P1","clinical-materials/Clinical_Implementation_Spine/","Faculty/resident track"],
["RSSM Master v11 (newest)","Teaching Spine","RSS Framework","DOCX","RSSM model","Resident→Faculty","Adv","Reference","High","Exists","P0","rssm-manual/RSSM_Master_v11.docx","Canonical; archive v10"],
["RSSM publication set (journal article, appendices)","Teaching Spine","RSS Publication","MD/DOCX","RSSM model","Faculty","Adv","Reference","Med","Exists","P3","rssm-manual/publication/","Keep"],
["RSSM Master v10 (stray copies)","Archive","Stale Version","DOCX/MD","RSSM model","—","—","—","Low","Archive","P3","~/Downloads + rssm-manual","Move to 99_Archive"],
["Psychoed Library (patient/family/infographics/journey)","Patient/Family Ed","Psychoeducation","Mixed","Patient education","Patient/Family","Intro","Reference","Med","Exists","P3","psychoed-library/ (processed)","Read-only ref in 10_"],
["Psychoed Library — Raw_Records (1,339)","Archive","Raw Export","MD","Raw export","—","—","—","Low","Archive","P3","psychoed-library/Raw_Records/","Archive after verifying processed"],
["Post-Discharge Kit (bundles by dx, perinatal, crisis)","Patient/Family Ed","Discharge Bundles","Mixed","Disposition","Patient/Family","Intro","Reference","Med","Exists","P3","post-discharge-kit/","Link to disposition"],
["Population Adaptations (Forensic, Refugee, IDD, Perinatal)","Tracks","Population-Specific","MD/PNG","Special populations","Resident→Faculty","Adv","Reference","Med","Exists","P3","population-adaptations/","Multi-track seed"],
["Depression First Steps; SSRI First 6 Weeks (patient)","Patient/Family Ed","Psychoeducation","HTML","Depression/SSRI","Patient","Intro","15m","Med","Exists","P3","~/Depression_First_Steps_PAT.html","Link in 10_"],
["Book + Podcast Library seed catalog (verified)","Media","Bibliotherapy","MD","Books/podcasts","Patient/Clinician","Intro","Reference","Med","Exists","P3","~/Library_Plan_and_Audit_Roadmap.md","Build Library v1"],
["Shelf High-Yield review guide","Exam Prep","Shelf/NBME","MD","Exam prep","MS3","Inter","2h","High","Exists","P2","14_Tracks/MS3/Student_Ready_Pack/07_shelf_guide/","Filled by Codex shelf review guide"],
["OSCE Station set (6 stations + checklists)","Exam Prep","OSCE","MD","OSCE","MS3","Inter","2h","High","Exists","P2","14_Tracks/MS3/Student_Ready_Pack/06_osce_cases/","Filled by Codex 6-station OSCE set"],
["Reflection + PIF prompt set (interactive)","Clinical Skills","Reflection/PIF","HTML","Professional identity","MS3","Intro","10m each","Med","Exists","P2","02_Clinical_Skills/Reflection_PIF/","BUILT: 6 weekly + ethics prompts, journaling"],
["MS3 Student Ready Pack (Codex — 15 files)","Tracks","MS3 Track","MD","Orientation/skills/OSCE/shelf/cases","MS3","Intro→Inter","Varies","High","Exists","P1","14_Tracks/MS3/Student_Ready_Pack/","Adopted from Codex; pairs with interactive tools"],
["Exhaustive file census + duplicate log (Codex)","Planning","Census/Data","CSV/JSON","Inventory","Faculty","—","Reference","High","Exists","P0","00_START_HERE/_audit-census-codex/","Inventory of record: 11,700 files / 2,785 dup groups"],
["Interview & MSE Pocket Guide (Codex)","Clinical Skills","Interviewing","MD","Interview/MSE","MS3","Intro","15m","High","Exists","P1","…/Student_Ready_Pack/02_pocket_guides/","Reference companion to MSE module"],
["MS3 Orientation Packet (Codex)","Orientation","Logistics","MD","Unit orientation","MS3","Intro","20m","High","Exists","P1","…/Student_Ready_Pack/01_orientation/","Closes orientation gap"],
["Clinical Warm Design System","Faculty","Design/Style","HTML","Styling","Faculty","—","Reference","Low","Exists","P3","~/clinical-warm-design-system.html","Apply to public mirror"],
]
header_fill = PatternFill("solid", start_color="1F4E5F")
header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
base_font = Font(name="Arial", size=10)
wrap = Alignment(vertical="top", wrap_text=True)
center = Alignment(vertical="top", horizontal="center", wrap_text=True)
thin = Side(style="thin", color="D9D9D9"); border = Border(left=thin,right=thin,top=thin,bottom=thin)
status_fill = {"Exists":"E2EFDA","Revise":"FFF2CC","Expand":"DDEBF7","Create":"FCE4D6","Merge":"FFF2CC","Archive":"F2F2F2"}
prio_fill = {"P0":"C00000","P1":"ED7D31","P2":"FFC000","P3":"A6A6A6"}
ws.append(cols)
for r in R: ws.append(r)
for c in range(1, len(cols)+1):
    cell = ws.cell(row=1, column=c); cell.fill=header_fill; cell.font=header_font
    cell.alignment=Alignment(vertical="center",horizontal="center",wrap_text=True); cell.border=border
status_idx=cols.index("Current Status")+1; prio_idx=cols.index("Priority")+1
center_cols={status_idx,prio_idx,cols.index("Difficulty")+1,cols.index("Est. Time")+1,cols.index("Clinical Relevance")+1,cols.index("Format")+1}
for row in range(2, ws.max_row+1):
    for c in range(1, len(cols)+1):
        cell=ws.cell(row=row,column=c); cell.font=base_font; cell.border=border
        cell.alignment = center if c in center_cols else wrap
    sc=ws.cell(row=row,column=status_idx)
    if sc.value in status_fill:
        sc.fill=PatternFill("solid",start_color=status_fill[sc.value]); sc.font=Font(name="Arial",size=10,bold=True)
    pc=ws.cell(row=row,column=prio_idx)
    if pc.value in prio_fill:
        pc.fill=PatternFill("solid",start_color=prio_fill[pc.value]); pc.font=Font(name="Arial",size=10,bold=True,color="FFFFFF")
widths=[40,18,20,12,20,16,11,12,12,13,9,34,32]
for i,w in enumerate(widths,start=1): ws.column_dimensions[get_column_letter(i)].width=w
ws.freeze_panes="A2"; ws.auto_filter.ref=f"A1:{get_column_letter(len(cols))}{ws.max_row}"; ws.row_dimensions[1].height=30
n=ws.max_row
s=wb.create_sheet("Summary")
s["A1"]="Psychiatry Clerkship Library — Index Summary"; s["A1"].font=Font(name="Arial",bold=True,size=13)
s["A2"]="Auto-computed from the Master Index sheet"; s["A2"].font=Font(name="Arial",italic=True,size=9,color="808080")
stat_rng=f"'Master Index'!$J$2:$J${n}"; cat_rng=f"'Master Index'!$B$2:$B${n}"; prio_rng=f"'Master Index'!$K$2:$K${n}"; a_rng=f"'Master Index'!$A$2:$A${n}"
s["A4"]="By Status"; s["A4"].font=Font(name="Arial",bold=True,size=11)
for i,st in enumerate(["Exists","Revise","Expand","Create","Merge","Archive"]):
    s.cell(row=5+i,column=1,value=st).font=base_font
    s.cell(row=5+i,column=2,value=f'=COUNTIF({stat_rng},"{st}")').font=base_font
s.cell(row=11,column=1,value="TOTAL").font=Font(name="Arial",bold=True)
s.cell(row=11,column=2,value=f"=COUNTA({a_rng})").font=Font(name="Arial",bold=True)
s["D4"]="By Priority"; s["D4"].font=Font(name="Arial",bold=True,size=11)
for i,p in enumerate(["P0","P1","P2","P3"]):
    s.cell(row=5+i,column=4,value=p).font=base_font
    s.cell(row=5+i,column=5,value=f'=COUNTIF({prio_rng},"{p}")').font=base_font
s["A14"]="By Category"; s["A14"].font=Font(name="Arial",bold=True,size=11)
for i,ct in enumerate(["Planning","Orientation","Evidence & Reading","Teaching Spine","Clinical Skills","Acute & Safety","Psychopharmacology","Family & Relational","Core Topics","Patient/Family Ed","Cases & Simulation","Exam Prep","Media","AI & Prompts","Faculty","Tracks","Archive"]):
    s.cell(row=15+i,column=1,value=ct).font=base_font
    s.cell(row=15+i,column=2,value=f'=COUNTIF({cat_rng},"{ct}")').font=base_font
s["D14"]="Status legend"; s["D14"].font=Font(name="Arial",bold=True,size=11)
for i,(k,v) in enumerate([("Exists","usable today"),("Revise","update/fix drift"),("Expand","extend existing"),("Create","net-new authoring"),("Merge","dedupe to canonical"),("Archive","retire")]):
    cc=s.cell(row=15+i,column=4,value=k); cc.fill=PatternFill("solid",start_color=status_fill[k]); cc.font=Font(name="Arial",bold=True,size=10)
    s.cell(row=15+i,column=5,value=v).font=base_font
for col,w in {"A":24,"B":10,"C":3,"D":14,"E":26}.items(): s.column_dimensions[col].width=w
import os
_here = os.path.dirname(os.path.abspath(__file__))
_lib = os.path.abspath(os.path.join(_here, "..", "..", ".."))   # repo root
out = os.environ.get("MASTER_INDEX_OUT", os.path.join(_lib, "_MASTER_INDEX.xlsx"))
wb.save(out); print("rows:", ws.max_row-1, "saved", out)
